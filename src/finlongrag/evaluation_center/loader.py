"""Test-set loading utilities."""

from __future__ import annotations

import csv
from pathlib import Path

from finlongrag.evaluation_center.schemas import TestSet, TestSetItem


def load_test_set(path: str | Path, *, test_set_id: str = "", name: str = "") -> TestSet:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        return load_test_set_from_xlsx(file_path, test_set_id=test_set_id, name=name)
    if suffix == ".csv":
        return load_test_set_from_csv(file_path, test_set_id=test_set_id, name=name)
    raise ValueError(f"unsupported test-set file type: {file_path.suffix}")


def load_test_set_from_csv(path: str | Path, *, test_set_id: str = "", name: str = "") -> TestSet:
    file_path = Path(path)
    items: list[TestSetItem] = []
    with file_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        for row in reader:
            question = _first(row, ["question", "query", "问题", fieldnames[0] if fieldnames else ""])
            if not question:
                continue
            items.append(
                TestSetItem(
                    question=question,
                    answer=_first(row, ["answer", "expected_answer", "答案", fieldnames[1] if len(fieldnames) > 1 else ""]),
                    source=_first(row, ["source", "expected_source", "来源", fieldnames[2] if len(fieldnames) > 2 else ""]),
                )
            )
    return TestSet(
        id=test_set_id,
        name=name or file_path.stem,
        sheet_names=[file_path.stem],
        items=items,
        file_name=file_path.name,
        count=len(items),
    )


def load_test_set_from_xlsx(path: str | Path, *, test_set_id: str = "", name: str = "") -> TestSet:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to load xlsx test sets") from exc

    file_path = Path(path)
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    items: list[TestSetItem] = []
    sheet_names = list(workbook.sheetnames)
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(cell or "").strip() for cell in rows[0]]
        for row in rows[1:]:
            row_map = {header[index]: _cell(row[index]) for index in range(min(len(header), len(row)))}
            question = _first(row_map, ["question", "query", "问题", header[0] if header else ""])
            if not question:
                continue
            items.append(
                TestSetItem(
                    question=question,
                    answer=_first(row_map, ["answer", "expected_answer", "答案", header[1] if len(header) > 1 else ""]),
                    source=_first(row_map, ["source", "expected_source", "来源", header[2] if len(header) > 2 else ""]),
                    metadata={"sheet": sheet_name},
                )
            )
    return TestSet(
        id=test_set_id,
        name=name or file_path.stem,
        sheet_names=sheet_names,
        items=items,
        file_name=file_path.name,
        count=len(items),
    )


def _first(row: dict[str, str], keys: list[str]) -> str:
    for key in keys:
        if key and key in row and str(row[key]).strip():
            return str(row[key]).strip()
    return ""


def _cell(value: object) -> str:
    return "" if value is None else str(value).strip()

