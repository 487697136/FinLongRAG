"""Document parsing utilities.

PDF files are converted with ``opendataloader-pdf`` (``pip install opendataloader-pdf``).
Other formats use simple built-in parsers.
"""

from __future__ import annotations

import csv
import html
import inspect
import json
import re
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

from finlongrag.core.config import Settings
from finlongrag.core.schema import Document, PageText

_PAGE_SEPARATOR = "\n\n---PAGE-%page-number%---\n\n"

_OPENDATALOADER_IMPORT_ERROR: ImportError | None = None
try:
    import opendataloader_pdf
except ImportError as exc:
    _OPENDATALOADER_IMPORT_ERROR = exc

    class _MissingOpenDataLoaderPdf:
        @staticmethod
        def convert(*_args: Any, **_kwargs: Any) -> None:
            raise RuntimeError(
                "opendataloader-pdf is required to parse PDF files. "
                "Install it with `pip install opendataloader-pdf`."
            ) from _OPENDATALOADER_IMPORT_ERROR

    opendataloader_pdf = _MissingOpenDataLoaderPdf()


def parse_document(
    document: Document,
    *,
    settings: Settings | None = None,
) -> tuple[Document, list[PageText]]:
    path = document.path_obj
    suffix = path.suffix.lower()

    if suffix == ".pdf":
        pages = _parse_pdf(path, document, settings)
    elif suffix in {".txt", ".md"}:
        pages = _parse_text(path, document)
    elif suffix in {".html", ".htm"}:
        pages = _parse_html(path, document)
    elif suffix == ".csv":
        pages = _parse_csv(path, document)
    elif suffix == ".json":
        pages = _parse_json(path, document)
    elif suffix == ".xlsx":
        pages = _parse_xlsx(path, document)
    else:
        raise ValueError(f"unsupported document type: {path}")

    raw_text = "\n\n".join(page.text for page in pages if page.text)
    parser_name = pages[0].metadata.get("parser", "finlongrag.parser") if pages else "finlongrag.parser"
    parsed = Document(
        doc_id=document.doc_id,
        domain=document.domain,
        title=document.title or _infer_title(raw_text, path.stem),
        path=document.path,
        raw_text=raw_text,
        metadata={**document.metadata, "page_count": len(pages), "parser": parser_name},
    )
    return parsed, pages


def _parse_pdf(path: Path, document: Document, settings: Settings | None) -> list[PageText]:
    convert_kwargs: dict = {
        "input_path": str(path.resolve()),
        "format": "markdown",
        "quiet": True,
        "markdown_with_html": True,
        "image_output": "off",
        "markdown_page_separator": _PAGE_SEPARATOR,
    }
    if settings and settings.pdf_hybrid and settings.pdf_hybrid.lower() not in {"off", "none", ""}:
        convert_kwargs["hybrid"] = settings.pdf_hybrid
        if settings.pdf_hybrid_url:
            convert_kwargs["hybrid_url"] = settings.pdf_hybrid_url

    conversion_error = _opendataloader_preflight_error()
    if conversion_error is None:
        with tempfile.TemporaryDirectory(prefix="finlongrag-pdf-") as tmp:
            output_dir = Path(tmp)
            try:
                convert_result = _convert_pdf(output_dir=output_dir, convert_kwargs=convert_kwargs)
                markdown_text = _extract_markdown_from_result(convert_result)
                if markdown_text:
                    return markdown_to_pages(markdown_text, document, path)

                md_path = _find_generated_file(output_dir, path.stem, suffix=".md")
                if md_path is not None:
                    return markdown_to_pages(md_path.read_text(encoding="utf-8"), document, path)

                text_pages = _parse_pdf_via_text_output(path, document, output_dir=output_dir, settings=settings)
                if text_pages:
                    return text_pages
            except Exception as exc:
                conversion_error = _format_pdf_conversion_error(exc)

    pymupdf_pages = _parse_pdf_via_pymupdf(path, document)
    if pymupdf_pages:
        return pymupdf_pages

    pypdf_pages = _parse_pdf_via_pypdf(path, document)
    if pypdf_pages:
        return pypdf_pages

    raise RuntimeError(
        f"failed to parse PDF {path.name}: {conversion_error or 'opendataloader-pdf produced no extractable text'}. "
        "No extractable text was found via PyMuPDF/pypdf fallback. "
        "If this is a scanned PDF, install Java 11+ for opendataloader-pdf or configure OCR support."
    )


def _convert_pdf(*, output_dir: Path, convert_kwargs: dict) -> Any:
    """Call opendataloader-pdf while tolerating minor version signature drift."""
    signature = inspect.signature(opendataloader_pdf.convert)
    accepted = set(signature.parameters)
    kwargs = {key: value for key, value in convert_kwargs.items() if key in accepted}
    if "output_dir" in accepted:
        kwargs["output_dir"] = str(output_dir)
        return opendataloader_pdf.convert(**kwargs)
    return opendataloader_pdf.convert(str(output_dir), **kwargs)


def _opendataloader_preflight_error() -> str | None:
    if _OPENDATALOADER_IMPORT_ERROR is not None:
        return "opendataloader-pdf is not installed"
    java_major = _java_major_version()
    if java_major is None:
        return "Java runtime was not found; install Java 11+ to use opendataloader-pdf"
    if java_major < 11:
        return f"Java {java_major} is too old for opendataloader-pdf; install Java 11+"
    return None


def _java_major_version() -> int | None:
    java = shutil.which("java")
    if not java:
        return None
    try:
        result = subprocess.run(
            [java, "-version"],
            capture_output=True,
            text=True,
            timeout=5,
            check=False,
        )
    except Exception:
        return None
    output = "\n".join(part for part in (result.stderr, result.stdout) if part)
    match = re.search(r'version\s+"(?P<version>\d+)(?:\.(?P<minor>\d+))?', output)
    if not match:
        return None
    major = int(match.group("version"))
    if major == 1 and match.group("minor"):
        return int(match.group("minor"))
    return major


def _format_pdf_conversion_error(exc: Exception) -> str:
    if isinstance(exc, subprocess.CalledProcessError):
        stderr = (exc.stderr or "").strip()
        if "UnsupportedClassVersionError" in stderr or "class file version 55" in stderr:
            return "Java runtime is too old for opendataloader-pdf; install Java 11+"
        if stderr:
            return f"opendataloader-pdf failed: {stderr[:500]}"
    return f"opendataloader-pdf failed: {exc.__class__.__name__}: {exc}"


def _extract_markdown_from_result(result: Any) -> str:
    """Best-effort extraction for newer opendataloader return payloads."""
    if isinstance(result, str):
        return result.strip()
    if isinstance(result, dict):
        for key in ("markdown", "md", "content", "text"):
            value = result.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
    return ""


def _find_generated_file(output_dir: Path, stem: str, *, suffix: str) -> Path | None:
    preferred = output_dir / f"{stem}{suffix}"
    if preferred.exists():
        return preferred
    matches = sorted(output_dir.rglob(f"*{suffix}"))
    return matches[0] if matches else None


def _parse_pdf_via_text_output(
    path: Path,
    document: Document,
    *,
    output_dir: Path,
    settings: Settings | None,
) -> list[PageText]:
    text_kwargs: dict[str, Any] = {
        "input_path": str(path.resolve()),
        "format": "text",
        "quiet": True,
        "text_page_separator": _PAGE_SEPARATOR,
    }
    if settings and settings.pdf_hybrid and settings.pdf_hybrid.lower() not in {"off", "none", ""}:
        text_kwargs["hybrid"] = settings.pdf_hybrid
        if settings.pdf_hybrid_url:
            text_kwargs["hybrid_url"] = settings.pdf_hybrid_url
    _convert_pdf(output_dir=output_dir, convert_kwargs=text_kwargs)
    txt_path = _find_generated_file(output_dir, path.stem, suffix=".txt")
    if txt_path is None:
        return []
    text = txt_path.read_text(encoding="utf-8", errors="ignore")
    return _text_to_pages(text, document, path, parser_name="opendataloader-pdf:text")


def _parse_pdf_via_pypdf(path: Path, document: Document) -> list[PageText]:
    try:
        from pypdf import PdfReader
    except Exception:
        return []

    pages: list[PageText] = []
    try:
        reader = PdfReader(str(path))
    except Exception:
        return []
    for page_no, page in enumerate(reader.pages, start=1):
        text = _normalize_text(page.extract_text() or "")
        if not text:
            continue
        pages.append(_page(document, path, page_no, text, parser_name="pypdf"))
    return pages


def _parse_pdf_via_pymupdf(path: Path, document: Document) -> list[PageText]:
    try:
        import fitz
    except Exception:
        return []

    pages: list[PageText] = []
    try:
        pdf = fitz.open(str(path))
    except Exception:
        return []
    try:
        can_try_ocr = _has_tesseract()
        for page_no, page in enumerate(pdf, start=1):
            text = _normalize_text(page.get_text("text") or "")
            if not text and can_try_ocr and hasattr(page, "get_textpage_ocr"):
                try:
                    ocr_page = page.get_textpage_ocr(dpi=300, full=True)
                    text = _normalize_text(page.get_text("text", textpage=ocr_page) or "")
                except Exception:
                    text = ""
            if not text:
                continue
            parser_name = "pymupdf+ocr" if can_try_ocr else "pymupdf"
            pages.append(_page(document, path, page_no, text, parser_name=parser_name))
    finally:
        pdf.close()
    return pages


def _has_tesseract() -> bool:
    return bool(shutil.which("tesseract"))


def markdown_to_pages(markdown_text: str, document: Document, source_path: Path) -> list[PageText]:
    """Split OpenDataLoader markdown output into PageText (one entry per page)."""
    parts = re.split(r"---PAGE-(\d+)---", markdown_text)
    pages: list[PageText] = []

    if len(parts) < 3:
        text = _normalize_text(markdown_text)
        if text:
            pages.append(_page(document, source_path, 1, text))
        return pages

    leading = _normalize_text(parts[0])
    if leading:
        pages.append(_page(document, source_path, 1, leading))

    index = 1
    while index < len(parts) - 1:
        page_no = int(parts[index])
        body = _normalize_text(parts[index + 1])
        if body:
            pages.append(_page(document, source_path, page_no, body))
        index += 2

    return pages


def _text_to_pages(text: str, document: Document, source_path: Path, *, parser_name: str) -> list[PageText]:
    parts = re.split(r"---PAGE-(\d+)---", text)
    pages: list[PageText] = []
    if len(parts) < 3:
        normalized = _normalize_text(text)
        if normalized:
            pages.append(_page(document, source_path, 1, normalized, parser_name=parser_name))
        return pages

    leading = _normalize_text(parts[0])
    if leading:
        pages.append(_page(document, source_path, 1, leading, parser_name=parser_name))
    index = 1
    while index < len(parts) - 1:
        page_no = int(parts[index])
        body = _normalize_text(parts[index + 1])
        if body:
            pages.append(_page(document, source_path, page_no, body, parser_name=parser_name))
        index += 2
    return pages


def _page(document: Document, source_path: Path, page_no: int, text: str, *, parser_name: str = "opendataloader-pdf") -> PageText:
    return PageText(
        document.doc_id,
        document.domain,
        page_no,
        text,
        {"source": str(source_path), "parser": parser_name},
    )


def _parse_text(path: Path, document: Document) -> list[PageText]:
    text = _normalize_text(path.read_text(encoding="utf-8", errors="ignore"))
    return [PageText(document.doc_id, document.domain, 1, text, {"source": str(path), "parser": "text"})]


def _parse_html(path: Path, document: Document) -> list[PageText]:
    raw = path.read_text(encoding="utf-8", errors="ignore")
    raw = re.sub(r"(?is)<(script|style).*?>.*?</\1>", "", raw)
    raw = re.sub(r"(?is)<br\s*/?>", "\n", raw)
    raw = re.sub(r"(?is)</p>|</div>|</tr>|</h[1-6]>", "\n", raw)
    text = html.unescape(re.sub(r"(?is)<.*?>", "", raw))
    return [PageText(document.doc_id, document.domain, 1, _normalize_text(text), {"source": str(path), "parser": "html"})]


def _parse_csv(path: Path, document: Document) -> list[PageText]:
    rows: list[str] = []
    with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(" | ".join(str(cell).strip() for cell in row if str(cell).strip()))
    text = _normalize_text("\n".join(row for row in rows if row))
    return [PageText(document.doc_id, document.domain, 1, text, {"source": str(path), "parser": "csv", "format": "csv"})]


def _parse_json(path: Path, document: Document) -> list[PageText]:
    raw = path.read_text(encoding="utf-8-sig", errors="ignore")
    try:
        data = json.loads(raw)
        text = json.dumps(data, ensure_ascii=False, indent=2)
    except json.JSONDecodeError:
        text = raw
    return [PageText(document.doc_id, document.domain, 1, _normalize_text(text), {"source": str(path), "parser": "json", "format": "json"})]


def _parse_xlsx(path: Path, document: Document) -> list[PageText]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to parse xlsx files") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    pages: list[PageText] = []
    try:
        for index, sheet in enumerate(workbook.worksheets, start=1):
            lines: list[str] = [f"# Sheet: {sheet.title}"]
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell not in (None, "")]
                if cells:
                    lines.append(" | ".join(cells))
            pages.append(
                PageText(
                    document.doc_id,
                    document.domain,
                    index,
                    _normalize_text("\n".join(lines)),
                    {"source": str(path), "parser": "xlsx", "format": "xlsx", "sheet": sheet.title},
                )
            )
    finally:
        workbook.close()
    return pages


def _normalize_text(text: str) -> str:
    text = text.replace("\u3000", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _infer_title(text: str, default_title: str) -> str:
    for line in text.splitlines()[:20]:
        line = line.strip().lstrip("#").strip()
        if len(line) >= 4 and not re.fullmatch(r"(?:第\s*)?\d{1,4}\s*(?:页|page)?", line, flags=re.IGNORECASE):
            return line[:120]
    return default_title
